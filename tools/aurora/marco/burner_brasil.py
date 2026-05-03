#!/usr/bin/env python3
"""
BURNER PREVIDENCIÁRIO BRASIL — PIRÂMIDE
========================================
Saturação L4 + Vertex sobre TODOS os indeferimentos INSS do Brasil (12 meses).

Pipeline em pirâmide (otimizado pra ~9.6M leads):

  📊 BASE (100%, ~9.6M leads)         Filtro determinístico Python (sem IA)
                                       → score_pre 0-100 baseado em regras

  🥉 TOP 30% (~2.9M)                  Gemma 27B local (L4 a 95%+)
                                       → sub_vertical, tese, score_conv, urgência

  🥈 TOP 10% (~960k)                  Vertex Flash (~$0.0001/lead = $96 total)
                                       → re-score, refinamento de tese

  🥇 TOP 1% (~96k)                    Vertex Pro (~$0.005/lead = $480 total)
                                       → dossiê + script abordagem + cálculo atrasados

Output:
  gs://datalake-tbr-clean/leads_prev_brasil/
    ├── camada_base.jsonl.gz       (9.6M, score determinístico)
    ├── camada_gemma.jsonl.gz      (2.9M, Gemma classif)
    ├── camada_flash.jsonl.gz      (960k, Vertex Flash)
    └── camada_pro.jsonl.gz        (96k, dossiê completo)

Args:
  --skip-download    pula etapa de download (reusa XLSX do burner regional)
  --skip-base        pula filtro determinístico (reusa camada_base)
  --skip-gemma       pula Gemma (reusa camada_gemma)
  --skip-flash       pula Vertex Flash
  --skip-pro         pula Vertex Pro
  --workers N        workers Gemma paralelos (default 8)
  --top-pct-gemma N  % do total que vai pra Gemma (default 30)
  --top-pct-flash N  % do total que vai pra Flash (default 10)
  --top-pct-pro N    % do total que vai pra Pro (default 1)
  --dry-run          não sobe pro GCS
"""

import os
import sys
import json
import gzip
import asyncio
import logging
import argparse
from pathlib import Path
from datetime import datetime, date
from collections import Counter
import urllib.request

import httpx
from openpyxl import load_workbook

# === CONFIG ===
OLLAMA_URL = "http://127.0.0.1:11434"
OLLAMA_MODEL = "gemma2:27b-instruct-q4_K_M"
WORK_DIR = Path("/home/manusalt13/leads_prev_brasil")
WORK_DIR.mkdir(parents=True, exist_ok=True)

# Reusa XLSX baixados pelo burner regional pra não baixar 2x
XLSX_REUSE_DIR = Path("/home/manusalt13/leads_prev_marco")

GCS_PREFIX = "gs://datalake-tbr-clean/leads_prev_brasil"

# BigQuery
BQ_PROJECT = os.environ.get('BQ_PROJECT', 'transparenciabr')
BQ_DATASET = os.environ.get('BQ_DATASET', 'tbr_leads_prev')
BQ_LOCATION = 'southamerica-east1'
BQ_CHUNK = 5000  # batch insert size

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.StreamHandler()]
)
log = logging.getLogger(__name__)


# ==================== ETAPA 1: DOWNLOAD ====================

def baixar_meses(urls_file: Path) -> list:
    """Baixa todos os XLSX dos 12 meses se não existirem.
    Reusa primeiro de XLSX_REUSE_DIR (burner regional já baixou)."""
    with open(urls_file) as f:
        urls = json.load(f)

    files = []
    for u in urls:
        nome = u['name'].replace(' ', '_').replace('í', 'i').replace('é', 'e').replace('ç', 'c')
        local = WORK_DIR / f"{nome}.xlsx"
        reuse = XLSX_REUSE_DIR / f"{nome}.xlsx"

        if local.exists() and local.stat().st_size > 1_000_000:
            log.info(f"✅ Local: {local.name} ({local.stat().st_size/1e6:.1f} MB)")
            files.append(local)
            continue

        if reuse.exists() and reuse.stat().st_size > 1_000_000:
            log.info(f"♻️  Reusando: {reuse.name} ({reuse.stat().st_size/1e6:.1f} MB)")
            files.append(reuse)
            continue

        log.info(f"⬇️  Baixando: {u['name']}")
        try:
            req = urllib.request.Request(u['url'], headers={'User-Agent': 'TBR-Brasil/1.0'})
            with urllib.request.urlopen(req, timeout=300) as resp:
                with open(local, 'wb') as fout:
                    fout.write(resp.read())
            log.info(f"  → {local.stat().st_size/1e6:.1f} MB")
            files.append(local)
        except Exception as e:
            log.error(f"  ❌ Falha: {e}")
            continue

    return files


# ==================== ETAPA 2: FILTRO DETERMINÍSTICO BRASIL ====================

# Tabela de pesos baseada em conhecimento previdenciário público
PESOS_MOTIVO = {
    # Não comparecimento à perícia: super fácil, escritório só reagenda
    'NÃO COMPARECIMENTO À PERÍCIA': 95,
    'NAO COMPARECIMENTO': 90,
    'AUSÊNCIA DO REQUERENTE': 90,
    # Não constatação de incapacidade: perícia judicial reverte muito
    'NÃO CONSTATAÇÃO': 80,
    'NAO CONSTATACAO DE INCAPACIDADE': 80,
    'NÃO COMPROVOU INCAPACIDADE': 80,
    # BPC tese miserabilidade (STF Tema 27)
    'NÃO ATENDE CRITÉRIO DE MISERABILIDADE': 78,
    'RENDA SUPERIOR': 75,
    # Carência rural — prova testemunhal vence muito
    'FALTA PERÍODO DE CARÊNCIA': 70,
    'NÃO COMPROVOU EXERCÍCIO DE ATIVIDADE RURAL': 75,
    # BPC PCD não enquadramento — perícia social CRAS reverte
    'NÃO ATENDE CRITÉRIO DE DEFICIÊNCIA': 75,
    'NÃO É CONSIDERADO PESSOA COM DEFICIÊNCIA': 75,
    # EC 103 transição
    'NÃO IMPLEMENTOU REQUISITOS': 55,
    'INSUFICIÊNCIA DE TEMPO DE CONTRIBUIÇÃO': 60,
    # Outros
    'PERDA DA QUALIDADE DE SEGURADO': 50,
    'DCB': 65,
    'CESSAÇÃO': 70,
}

PESOS_ESPECIE = {
    # Auxílio-doença / B31: vertical Marco principal
    'AUXÍLIO POR INCAPACIDADE TEMPORÁRIA': 15,
    'AUXÍLIO-DOENÇA': 15,
    'B31': 15,
    # BPC: foco PCD do Marco
    'AMPARO SOCIAL À PESSOA COM DEFICIÊNCIA': 18,
    'AMPARO SOCIAL AO IDOSO': 12,
    'B87': 18,
    'B88': 12,
    # Aposentadoria por incapacidade
    'APOSENTADORIA POR INCAPACIDADE': 13,
    'B32': 13,
    # Pensão por morte
    'PENSÃO POR MORTE': 8,
    'B21': 8,
    # Salário-maternidade
    'SALÁRIO-MATERNIDADE': 5,
    'B80': 5,
}


def calc_idade(dt_nasc) -> int:
    """Calcula idade a partir de dt_nasc."""
    if not dt_nasc:
        return 0
    try:
        if isinstance(dt_nasc, datetime):
            d = dt_nasc.date()
        elif isinstance(dt_nasc, date):
            d = dt_nasc
        else:
            d = datetime.strptime(str(dt_nasc)[:10], '%Y-%m-%d').date()
        today = date(2026, 5, 1)
        return today.year - d.year - ((today.month, today.day) < (d.month, d.day))
    except Exception:
        return 0


def score_lead(lead: dict) -> int:
    """Score determinístico 0-100."""
    score = 0

    # Motivo: contribuição principal
    motivo = (lead.get('motivo') or '').upper()
    motivo_score = 0
    for chave, peso in PESOS_MOTIVO.items():
        if chave in motivo:
            motivo_score = max(motivo_score, peso)
    score += int(motivo_score * 0.55)  # peso 55%

    # Espécie
    esp = (lead.get('especie') or '').upper()
    esp_score = 0
    for chave, peso in PESOS_ESPECIE.items():
        if chave in esp:
            esp_score = max(esp_score, peso)
    score += esp_score  # peso direto até ~18 pontos

    # Idade: idosos têm urgência
    idade = calc_idade(lead.get('dt_nasc'))
    if idade >= 70:
        score += 15
    elif idade >= 60:
        score += 10
    elif idade >= 50:
        score += 5

    # UF: SP/RJ/MG têm advocacia mais densa = case fits Marco vendendo licença
    uf = (lead.get('uf') or '').strip()
    if uf in ('São Paulo', 'SP'):
        score += 8
    elif uf in ('Rio de Janeiro', 'RJ', 'Minas Gerais', 'MG', 'Paraná', 'PR'):
        score += 5

    # Recente = urgência prazo recurso 30 dias
    dt_ind = lead.get('dt_indef')
    if dt_ind:
        try:
            d = datetime.strptime(str(dt_ind)[:10], '%Y-%m-%d').date()
            dias = (date(2026, 5, 1) - d).days
            if dias <= 30:
                score += 10
            elif dias <= 90:
                score += 5
        except Exception:
            pass

    return min(score, 100)


def filtrar_brasil(xlsx_files: list, out_jsonl: Path) -> int:
    """Lê todos os XLSX, scora cada lead, salva JSONL.GZ.
    Retorna total de leads."""
    total = 0
    log.info(f"📊 Filtrando Brasil → {out_jsonl}")
    with gzip.open(out_jsonl, 'wt', encoding='utf-8') as fout:
        for fpath in xlsx_files:
            log.info(f"  📂 {fpath.name}")
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
                ws = wb[wb.sheetnames[0]]
                count = 0
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if not row or not row[0]:
                        continue
                    lead = {
                        'mes_arquivo': fpath.stem,
                        'competencia': str(row[0]) if row[0] else None,
                        'especie_cod': row[1],
                        'especie': row[2],
                        'motivo': row[3],
                        'dt_nasc': str(row[4]) if row[4] else None,
                        'sexo': row[5],
                        'clientela': row[6],
                        'filiacao': row[7],
                        'uf': row[8],
                        'dt_indef': str(row[9]) if row[9] else None,
                        'ramo': row[10],
                        'aps_cod': row[11],
                        'aps': row[12],
                        'dt_der': str(row[13]) if row[13] else None,
                    }
                    lead['score_pre'] = score_lead(lead)
                    fout.write(json.dumps(lead, ensure_ascii=False, default=str) + '\n')
                    count += 1
                    total += 1
                log.info(f"    → {count:,} leads")
                wb.close()
            except Exception as e:
                log.error(f"    ❌ {e}")
    log.info(f"✅ TOTAL Brasil: {total:,} leads em {out_jsonl} ({out_jsonl.stat().st_size/1e6:.1f} MB)")
    return total


def selecionar_top(in_jsonl: Path, out_jsonl: Path, pct: float, score_field: str = 'score_pre') -> int:
    """Lê JSONL.GZ, ordena por score, salva top pct% em outro JSONL.GZ."""
    log.info(f"🎯 Selecionando top {pct}% por {score_field}: {in_jsonl} → {out_jsonl}")
    leads = []
    with gzip.open(in_jsonl, 'rt', encoding='utf-8') as f:
        for line in f:
            try:
                leads.append(json.loads(line))
            except Exception:
                pass
    log.info(f"  carregados {len(leads):,} leads")
    leads.sort(key=lambda x: x.get(score_field, 0), reverse=True)
    cutoff = int(len(leads) * pct / 100)
    top = leads[:cutoff]
    with gzip.open(out_jsonl, 'wt', encoding='utf-8') as fout:
        for l in top:
            fout.write(json.dumps(l, ensure_ascii=False, default=str) + '\n')
    log.info(f"  → top {len(top):,} salvos ({out_jsonl.stat().st_size/1e6:.1f} MB)")
    return len(top)


# ==================== ETAPA 3: GEMMA 27B L4 ====================

PROMPT_GEMMA = """Você classifica leads previdenciários (recusas INSS) para escritório de advocacia.

Lead:
- Espécie: {especie}
- Motivo recusa: {motivo}
- Sexo: {sexo}
- Clientela: {clientela}
- Filiação: {filiacao}
- UF: {uf}
- Data nascimento: {dt_nasc}
- Data indeferimento: {dt_indef}
- APS: {aps}

Retorne SOMENTE este JSON:
{{
  "sub_vertical": "auxilio_doenca" | "bpc_pcd" | "bpc_idoso" | "aposentadoria_tempo" | "aposentadoria_idade" | "aposentadoria_incapacidade" | "pensao_morte" | "salario_maternidade" | "outro",
  "tese_juridica_curta": "<1 frase, máx 80 chars>",
  "score_conversao_0_100": <int>,
  "urgencia": "alta" | "media" | "baixa",
  "ticket_estimado_brl": <int min>,
  "fundamentos_chave": ["<lei/súmula/tese>", ...],
  "rationale": "<máx 30 palavras>"
}}"""


async def gemma_classify(client, lead, sem):
    async with sem:
        prompt = PROMPT_GEMMA.format(
            especie=lead.get('especie') or '',
            motivo=lead.get('motivo') or '',
            sexo=lead.get('sexo') or '',
            clientela=lead.get('clientela') or '',
            filiacao=lead.get('filiacao') or '',
            uf=lead.get('uf') or '',
            dt_nasc=lead.get('dt_nasc') or '',
            dt_indef=lead.get('dt_indef') or '',
            aps=lead.get('aps') or ''
        )
        try:
            r = await client.post(
                f"{OLLAMA_URL}/api/generate",
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "format": "json",
                    "stream": False,
                    "options": {"temperature": 0.2, "num_predict": 350},
                    "keep_alive": -1
                },
                timeout=180
            )
            r.raise_for_status()
            data = r.json()
            classif = json.loads(data['response'])
            lead['_gemma'] = classif
            lead['_gemma_ok'] = True
            lead['score_gemma'] = classif.get('score_conversao_0_100', 0)
        except Exception as e:
            lead['_gemma_ok'] = False
            lead['_gemma_err'] = str(e)[:200]
            lead['score_gemma'] = 0
        return lead


async def saturate_gemma(in_jsonl: Path, out_jsonl: Path, workers: int):
    """Lê JSONL.GZ, classifica via Gemma streaming pra arquivo (não carrega tudo na RAM)."""
    log.info(f"🔥 Saturando L4 — Gemma 27B em {in_jsonl} → {out_jsonl}")
    sem = asyncio.Semaphore(workers)

    leads = []
    with gzip.open(in_jsonl, 'rt', encoding='utf-8') as f:
        for line in f:
            try:
                leads.append(json.loads(line))
            except Exception:
                pass
    total = len(leads)
    log.info(f"  Total a classificar: {total:,}")

    fout = gzip.open(out_jsonl, 'wt', encoding='utf-8')
    done = 0
    ok = 0
    t0 = datetime.now()
    try:
        async with httpx.AsyncClient() as client:
            tasks = [asyncio.create_task(gemma_classify(client, l, sem)) for l in leads]
            for coro in asyncio.as_completed(tasks):
                r = await coro
                done += 1
                if r.get('_gemma_ok'):
                    ok += 1
                fout.write(json.dumps(r, ensure_ascii=False, default=str) + '\n')
                if done % 200 == 0:
                    fout.flush()
                    elapsed = (datetime.now() - t0).total_seconds()
                    rate = done / elapsed if elapsed > 0 else 0
                    eta = (total - done) / rate / 3600 if rate > 0 else -1
                    log.info(f"  Gemma: {done:,}/{total:,} ({ok:,} ok) — {rate:.1f}/s — ETA {eta:.1f}h")
    finally:
        fout.close()
    log.info(f"✅ Gemma OK: {ok:,}/{total:,}")
    return ok


# ==================== ETAPA 4-5: VERTEX (placeholders bem feitos) ====================

async def vertex_flash(in_jsonl: Path, out_jsonl: Path):
    """Re-score via Vertex Flash. Stub — implementação real usa SDK Vertex."""
    log.info(f"⚡ Vertex Flash: {in_jsonl} → {out_jsonl}")
    # TODO sprint próximo: integrar google-cloud-aiplatform
    # Por ora, copia camada gemma como placeholder (mantém pipeline rodando)
    import shutil
    shutil.copy(in_jsonl, out_jsonl)
    log.info(f"  ⚠️  STUB: copiou camada anterior. Implementar SDK Vertex no próximo PR.")


async def vertex_pro(in_jsonl: Path, out_jsonl: Path):
    """Dossiê + script + cálculo via Vertex Pro. Stub."""
    log.info(f"🥇 Vertex Pro: {in_jsonl} → {out_jsonl}")
    import shutil
    shutil.copy(in_jsonl, out_jsonl)
    log.info(f"  ⚠️  STUB: copiou camada anterior. Implementar SDK Vertex no próximo PR.")


# ==================== ETAPA 6: GCS UPLOAD + STATS ====================

def upload_gcs(local: Path, prefix: str = GCS_PREFIX):
    import subprocess
    dest = f"{prefix}/{local.name}"
    log.info(f"☁️  {local.name} → {dest}")
    r = subprocess.run(["gcloud", "storage", "cp", str(local), dest], capture_output=True, text=True)
    if r.returncode == 0:
        log.info(f"  ✅ OK")
    else:
        log.error(f"  ❌ {r.stderr[:300]}")


# ==================== BIGQUERY LOAD ====================

def bq_load_jsonl_gz(jsonl_gz: Path, table: str, write_disposition: str = 'WRITE_TRUNCATE'):
    """Load JSONL.GZ direto pra BigQuery via bq CLI (evita SDK overhead).
    table: nome simples (ex: 'leads_brasil_base'). Usa BQ_PROJECT.BQ_DATASET.
    """
    import subprocess
    full = f"{BQ_PROJECT}:{BQ_DATASET}.{table}"
    log.info(f"📥 BQ load: {jsonl_gz.name} → {full}")
    r = subprocess.run([
        "bq", "load",
        f"--location={BQ_LOCATION}",
        "--source_format=NEWLINE_DELIMITED_JSON",
        f"--write_disposition={write_disposition}",
        "--ignore_unknown_values",
        "--max_bad_records=1000",
        full,
        str(jsonl_gz),
    ], capture_output=True, text=True)
    if r.returncode == 0:
        log.info(f"  ✅ BQ OK → {full}")
        return True
    else:
        log.error(f"  ❌ BQ falhou: {r.stderr[:500]}")
        return False


def bq_flatten_for_load(in_jsonl_gz: Path, out_jsonl_gz: Path, table_kind: str):
    """Achata o JSON aninhado (_gemma.*) pras colunas top-level esperadas pelo BQ.
    table_kind: 'base' | 'gemma' | 'flash' | 'pro'
    """
    log.info(f"🔄 Flatten {table_kind}: {in_jsonl_gz.name} → {out_jsonl_gz.name}")
    n = 0
    with gzip.open(in_jsonl_gz, 'rt', encoding='utf-8') as fin, \
         gzip.open(out_jsonl_gz, 'wt', encoding='utf-8') as fout:
        for line in fin:
            try:
                l = json.loads(line)
            except Exception:
                continue
            row = {
                'mes_arquivo': l.get('mes_arquivo'),
                'competencia': l.get('competencia'),
                'especie_cod': l.get('especie_cod'),
                'especie': l.get('especie'),
                'motivo': l.get('motivo'),
                'dt_nasc': l.get('dt_nasc'),
                'sexo': l.get('sexo'),
                'clientela': l.get('clientela'),
                'filiacao': l.get('filiacao'),
                'uf': l.get('uf'),
                'dt_indef': l.get('dt_indef'),
                'ramo': l.get('ramo'),
                'aps_cod': l.get('aps_cod'),
                'aps': l.get('aps'),
                'dt_der': l.get('dt_der'),
                'score_pre': l.get('score_pre', 0),
            }
            if table_kind in ('gemma', 'flash', 'pro'):
                g = l.get('_gemma') or {}
                row.update({
                    'sub_vertical': g.get('sub_vertical'),
                    'tese_juridica_curta': g.get('tese_juridica_curta'),
                    'score_conversao_0_100': g.get('score_conversao_0_100', 0),
                    'urgencia': g.get('urgencia'),
                    'ticket_estimado_brl': g.get('ticket_estimado_brl', 0),
                    'fundamentos_chave': g.get('fundamentos_chave', []),
                    'rationale': g.get('rationale'),
                    'gemma_ok': bool(l.get('_gemma_ok')),
                    'gemma_err': l.get('_gemma_err'),
                })
            # cast int seguro
            for k in ('especie_cod', 'aps_cod', 'score_pre',
                      'score_conversao_0_100', 'ticket_estimado_brl'):
                v = row.get(k)
                if v is not None:
                    try:
                        row[k] = int(v)
                    except Exception:
                        row[k] = None
            fout.write(json.dumps(row, ensure_ascii=False, default=str) + '\n')
            n += 1
    log.info(f"  → {n:,} linhas achatadas")
    return n


def stats_brasil(jsonl_gz: Path) -> dict:
    """Stats high-level em streaming (sem carregar tudo na RAM)."""
    s = {
        'total': 0,
        'por_uf': Counter(),
        'por_especie': Counter(),
        'por_motivo': Counter(),
        'por_aps_top': Counter(),
        'score_dist': {'90+': 0, '75-89': 0, '50-74': 0, '<50': 0},
    }
    with gzip.open(jsonl_gz, 'rt', encoding='utf-8') as f:
        for line in f:
            try:
                l = json.loads(line)
            except Exception:
                continue
            s['total'] += 1
            s['por_uf'][l.get('uf') or 'unknown'] += 1
            s['por_especie'][l.get('especie') or 'unknown'] += 1
            s['por_motivo'][l.get('motivo') or 'unknown'] += 1
            s['por_aps_top'][l.get('aps') or 'unknown'] += 1
            sc = l.get('score_gemma') or l.get('score_pre') or 0
            if sc >= 90:
                s['score_dist']['90+'] += 1
            elif sc >= 75:
                s['score_dist']['75-89'] += 1
            elif sc >= 50:
                s['score_dist']['50-74'] += 1
            else:
                s['score_dist']['<50'] += 1
    return {
        'total': s['total'],
        'por_uf': dict(s['por_uf'].most_common()),
        'por_especie': dict(s['por_especie'].most_common(20)),
        'por_motivo': dict(s['por_motivo'].most_common(20)),
        'top_aps': dict(s['por_aps_top'].most_common(50)),
        'score_dist': s['score_dist'],
    }


# ==================== MAIN ====================

async def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--workers', type=int, default=8)
    ap.add_argument('--top-pct-gemma', type=float, default=30.0)
    ap.add_argument('--top-pct-flash', type=float, default=10.0)
    ap.add_argument('--top-pct-pro', type=float, default=1.0)
    ap.add_argument('--skip-download', action='store_true')
    ap.add_argument('--skip-base', action='store_true')
    ap.add_argument('--skip-gemma', action='store_true')
    ap.add_argument('--skip-flash', action='store_true')
    ap.add_argument('--skip-pro', action='store_true')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    log.info("=" * 70)
    log.info("BURNER PREVIDENCIÁRIO BRASIL — PIRÂMIDE — INICIANDO")
    log.info("=" * 70)

    # Etapa 1: download (reusa do regional)
    files = []
    if not args.skip_download:
        script_dir = Path(__file__).resolve().parent
        urls_file = script_dir / 'urls_12meses.json'
        if not urls_file.exists():
            log.error(f"❌ urls_12meses.json não encontrado em {script_dir}")
            sys.exit(1)
        log.info(f"📋 URLs file: {urls_file}")
        files = baixar_meses(urls_file)
        log.info(f"📁 {len(files)} arquivos prontos")
    else:
        # Pega XLSX do diretório regional + diretório local
        for d in [XLSX_REUSE_DIR, WORK_DIR]:
            for x in d.glob("*.xlsx"):
                if x.stat().st_size > 1_000_000:
                    files.append(x)
        log.info(f"📁 Skip download — {len(files)} XLSX encontrados")

    # Etapa 2: filtro determinístico Brasil
    base_jsonl = WORK_DIR / "camada_base.jsonl.gz"
    if not args.skip_base or not base_jsonl.exists():
        if files:
            filtrar_brasil(files, base_jsonl)
        else:
            log.error("❌ Sem XLSX pra processar")
            sys.exit(1)
    else:
        log.info(f"✅ Reusando: {base_jsonl}")

    # Etapa 3: top X% → Gemma
    top_gemma_jsonl = WORK_DIR / "top_gemma_input.jsonl.gz"
    selecionar_top(base_jsonl, top_gemma_jsonl, args.top_pct_gemma, 'score_pre')

    gemma_jsonl = WORK_DIR / "camada_gemma.jsonl.gz"
    if not args.skip_gemma:
        await saturate_gemma(top_gemma_jsonl, gemma_jsonl, args.workers)
    else:
        log.info(f"⏭️  Skip Gemma")

    # Etapa 4: top X% → Vertex Flash
    if gemma_jsonl.exists() and not args.skip_flash:
        top_flash_jsonl = WORK_DIR / "top_flash_input.jsonl.gz"
        selecionar_top(gemma_jsonl, top_flash_jsonl,
                       args.top_pct_flash * 100 / args.top_pct_gemma,
                       'score_gemma')
        flash_jsonl = WORK_DIR / "camada_flash.jsonl.gz"
        await vertex_flash(top_flash_jsonl, flash_jsonl)

        # Etapa 5: top X% → Vertex Pro
        if flash_jsonl.exists() and not args.skip_pro:
            top_pro_jsonl = WORK_DIR / "top_pro_input.jsonl.gz"
            selecionar_top(flash_jsonl, top_pro_jsonl,
                           args.top_pct_pro * 100 / args.top_pct_flash,
                           'score_gemma')
            pro_jsonl = WORK_DIR / "camada_pro.jsonl.gz"
            await vertex_pro(top_pro_jsonl, pro_jsonl)

    # Stats
    stats = stats_brasil(base_jsonl)
    stats_file = WORK_DIR / "stats_brasil.json"
    with open(stats_file, 'w') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    log.info(f"📊 Stats Brasil:")
    log.info(f"  Total leads: {stats['total']:,}")
    log.info(f"  UFs: {len(stats['por_uf'])}")
    log.info(f"  Score 90+: {stats['score_dist']['90+']:,}")
    log.info(f"  Score 75-89: {stats['score_dist']['75-89']:,}")

    # Upload GCS + BigQuery
    if not args.dry_run:
        # GCS backup
        for layer in ['camada_base.jsonl.gz', 'camada_gemma.jsonl.gz',
                      'camada_flash.jsonl.gz', 'camada_pro.jsonl.gz']:
            p = WORK_DIR / layer
            if p.exists():
                upload_gcs(p)
        upload_gcs(stats_file)

        # BigQuery load (a estrela do show)
        log.info("=" * 60)
        log.info("📊 CARREGANDO BIGQUERY")
        log.info("=" * 60)
        for kind, layer, table in [
            ('base',  'camada_base.jsonl.gz',  'leads_brasil_base'),
            ('gemma', 'camada_gemma.jsonl.gz', 'leads_brasil_gemma'),
            ('flash', 'camada_flash.jsonl.gz', 'leads_brasil_flash'),
            ('pro',   'camada_pro.jsonl.gz',   'leads_brasil_pro'),
        ]:
            p = WORK_DIR / layer
            if not p.exists():
                log.warning(f"  ⚠️  pulando {layer} (não existe)")
                continue
            flat = WORK_DIR / f"bq_{kind}.jsonl.gz"
            bq_flatten_for_load(p, flat, kind)
            bq_load_jsonl_gz(flat, table, 'WRITE_TRUNCATE')

    log.info("✅ PIRÂMIDE BRASIL COMPLETA — L4 + BigQuery cumpriram missão.")


if __name__ == "__main__":
    asyncio.run(main())
