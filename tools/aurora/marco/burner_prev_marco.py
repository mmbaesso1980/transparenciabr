#!/usr/bin/env python3
"""
BURNER PREVIDENCIÁRIO MARCO CARPES
==================================
Saturação L4 + Vertex pra reunião segunda-feira.

Pipeline:
1. Baixa 12 meses de microdados INSS indeferidos
2. Filtra região Carpes Mathias (raio 200km Pirassununga + Valinhos)
3. Gemma 27B local: classifica cada lead em
   - tese_juridica
   - sub_vertical (auxilio_doenca, BPC_PCD, BPC_idoso, aposentadoria_tempo, aposentadoria_idade)
   - score_conversao (0-100)
   - urgencia (alta/media/baixa)
4. Vertex Flash: re-score nos top 30%
5. Vertex Pro: enriquece top 10% com dossiê + script abordagem + cálculo atrasados
6. Salva em gs://datalake-tbr-clean/leads_prev_marco/

Args: --dry-run pra testar sem GCS upload
"""

import os
import sys
import json
import asyncio
import logging
import argparse
from pathlib import Path
from datetime import datetime
from collections import Counter
import urllib.request
import urllib.parse

import httpx
from openpyxl import load_workbook

# === CONFIG ===
OLLAMA_URL = "http://127.0.0.1:11434"
OLLAMA_MODEL = "gemma2:27b-instruct-q4_K_M"
WORK_DIR = Path("/home/manusalt13/leads_prev_marco")
WORK_DIR.mkdir(parents=True, exist_ok=True)

# Região Carpes Mathias (raio ~200km Pirassununga+Valinhos)
ALVOS_APS = [
    'VALINHOS', 'VINHEDO', 'ITATIBA', 'CAMPINAS', 'PIRASSUNUNGA',
    'LIMEIRA', 'RIO CLARO', 'JUNDIAI', 'BRAGANCA', 'PORTO FERREIRA',
    'SAO JOAO DA BOA VISTA', 'AMERICANA', 'SUMARE', 'MOGI MIRIM',
    'MOGI GUACU', 'ARARAS', 'LEME', 'SANTA BARBARA', 'NOVA ODESSA',
    'INDAIATUBA', 'HORTOLANDIA', 'PAULINIA', 'COSMOPOLIS', 'ARTUR NOGUEIRA'
]

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler('/var/log/tbr/burner_prev_marco.log'),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)


def baixar_meses(urls_file: Path) -> list:
    """Baixa todos os XLSX dos 12 meses se não existirem."""
    with open(urls_file) as f:
        urls = json.load(f)

    files = []
    for u in urls:
        nome = u['name'].replace(' ', '_').replace('í', 'i').replace('é', 'e').replace('ç', 'c')
        dest = WORK_DIR / f"{nome}.xlsx"
        if dest.exists() and dest.stat().st_size > 1_000_000:
            log.info(f"✅ Já tenho: {dest.name} ({dest.stat().st_size/1e6:.1f} MB)")
        else:
            log.info(f"⬇️  Baixando: {u['name']}")
            try:
                req = urllib.request.Request(u['url'], headers={'User-Agent': 'TBR-Carpes/1.0'})
                with urllib.request.urlopen(req, timeout=180) as resp:
                    with open(dest, 'wb') as fout:
                        fout.write(resp.read())
                log.info(f"  → {dest.stat().st_size/1e6:.1f} MB")
            except Exception as e:
                log.error(f"  ❌ Falha: {e}")
                continue
        files.append(dest)
    return files


def filtrar_alvo(xlsx_files: list) -> list:
    """Filtra leads da região Marco Carpes em todos os meses."""
    leads = []
    for fpath in xlsx_files:
        log.info(f"📊 Lendo: {fpath.name}")
        try:
            wb = load_workbook(fpath, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            count = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not row or not row[0]:
                    continue
                if row[8] != 'São Paulo':
                    continue
                aps = str(row[12] or '').upper()
                for alvo in ALVOS_APS:
                    if alvo in aps:
                        leads.append({
                            'mes_arquivo': fpath.stem,
                            'competencia': row[0],
                            'especie_cod': row[1],
                            'especie': row[2],
                            'motivo': row[3],
                            'dt_nasc': str(row[4]) if row[4] else None,
                            'sexo': row[5],
                            'clientela': row[6],
                            'filiacao': row[7],
                            'uf': row[8],
                            'dt_indef': str(row[9]) if row[9] else None,
                            'ramo': row[10],
                            'aps_cod': row[11],
                            'aps': row[12],
                            'dt_der': str(row[13]) if row[13] else None,
                            'aps_match': alvo
                        })
                        count += 1
                        break
            log.info(f"  → {count:,} leads alvo")
            wb.close()
        except Exception as e:
            log.error(f"  ❌ {e}")
    log.info(f"✅ TOTAL leads região Carpes (12 meses): {len(leads):,}")
    return leads


PROMPT_GEMMA = """Você é classificador de leads previdenciários para escritório de advocacia que atua em recusas do INSS.

Lead INSS:
- Espécie: {especie}
- Motivo recusa: {motivo}
- Sexo: {sexo}
- Clientela: {clientela}
- Filiação: {filiacao}
- Data nascimento: {dt_nasc}
- Data indeferimento: {dt_indef}
- APS: {aps}
- Ramo: {ramo}

Retorne SOMENTE este JSON:
{{
  "sub_vertical": "auxilio_doenca" | "bpc_pcd" | "bpc_idoso" | "aposentadoria_tempo" | "aposentadoria_idade" | "pensao_morte" | "salario_maternidade" | "outro",
  "tese_juridica_curta": "<1 frase, máx 80 chars>",
  "score_conversao_0_100": <int>,
  "urgencia": "alta" | "media" | "baixa",
  "ticket_estimado_brl": <int min>,
  "fundamentos_chave": ["<lei/súmula/tese>", ...],
  "rationale": "<máx 30 palavras>"
}}

Regras de score:
- "Não comparecimento perícia" → score 90+ (super fácil reagendar)
- "Não constatação incapacidade" → score 75-85 (perícia judicial reverte muito)
- "Não enquadramento art 20 §3" → score 70-80 (BPC tese miserabilidade STF Tema 27)
- "Falta requisitos transição EC 103" → score 50-65 (revisão complexa)
- "Falta período carência rural" → score 60-75 (prova testemunhal vence)
- "Não atende critério deficiência BPC" → score 70-85 (perícia social CRAS)
- Cliente >55 anos → +10 score (urgência)
- Cliente >70 anos → +15 score
"""


async def gemma_classify(client: httpx.AsyncClient, lead: dict, sem: asyncio.Semaphore) -> dict:
    """Classifica 1 lead via Ollama Gemma 27B."""
    async with sem:
        prompt = PROMPT_GEMMA.format(**lead)
        try:
            r = await client.post(
                f"{OLLAMA_URL}/api/generate",
                json={
                    "model": OLLAMA_MODEL,
                    "prompt": prompt,
                    "format": "json",
                    "stream": False,
                    "options": {"temperature": 0.2, "num_predict": 400},
                    "keep_alive": -1
                },
                timeout=120
            )
            r.raise_for_status()
            data = r.json()
            classif = json.loads(data['response'])
            lead['_gemma'] = classif
            lead['_gemma_ok'] = True
            return lead
        except Exception as e:
            lead['_gemma_ok'] = False
            lead['_gemma_err'] = str(e)[:200]
            return lead


async def saturate_l4(leads: list, workers: int = 6) -> list:
    """Satura L4 com classificação Gemma — workers paralelos."""
    log.info(f"🔥 Saturando L4 com {workers} workers Gemma 27B em {len(leads):,} leads")
    sem = asyncio.Semaphore(workers)
    async with httpx.AsyncClient() as client:
        tasks = [gemma_classify(client, lead, sem) for lead in leads]
        done = 0
        results = []
        for coro in asyncio.as_completed(tasks):
            r = await coro
            done += 1
            results.append(r)
            if done % 100 == 0:
                ok = sum(1 for x in results if x.get('_gemma_ok'))
                log.info(f"  Progresso: {done:,}/{len(leads):,} ({ok:,} ok)")
        return results


def to_jsonl(leads: list, out_file: Path):
    """Salva como JSONL."""
    with open(out_file, 'w') as f:
        for l in leads:
            f.write(json.dumps(l, ensure_ascii=False, default=str) + '\n')
    log.info(f"💾 Salvo: {out_file} ({len(leads):,} linhas)")


def upload_gcs(jsonl_file: Path):
    """Upload pro GCS."""
    import subprocess
    dest = f"gs://datalake-tbr-clean/leads_prev_marco/{jsonl_file.name}"
    log.info(f"☁️  Subindo: {dest}")
    r = subprocess.run(
        ["gcloud", "storage", "cp", str(jsonl_file), dest],
        capture_output=True, text=True
    )
    if r.returncode == 0:
        log.info(f"  ✅ OK")
    else:
        log.error(f"  ❌ {r.stderr[:300]}")


def stats_summary(leads: list) -> dict:
    """Resumo estatístico pro Marco."""
    s = {
        'total': len(leads),
        'classificados_ok': sum(1 for l in leads if l.get('_gemma_ok')),
        'por_aps': dict(Counter(l['aps'] for l in leads).most_common()),
        'por_especie': dict(Counter(l['especie'] for l in leads).most_common()),
        'por_motivo': dict(Counter(l['motivo'] for l in leads).most_common(15)),
        'por_subvertical': dict(Counter(
            l.get('_gemma', {}).get('sub_vertical', 'unknown') for l in leads if l.get('_gemma_ok')
        ).most_common()),
        'score_distribution': {
            'altissimo_90+': sum(1 for l in leads if l.get('_gemma', {}).get('score_conversao_0_100', 0) >= 90),
            'alto_75_89': sum(1 for l in leads if 75 <= l.get('_gemma', {}).get('score_conversao_0_100', 0) < 90),
            'medio_50_74': sum(1 for l in leads if 50 <= l.get('_gemma', {}).get('score_conversao_0_100', 0) < 75),
            'baixo_lt50': sum(1 for l in leads if l.get('_gemma', {}).get('score_conversao_0_100', 0) < 50),
        },
        'urgencia': dict(Counter(
            l.get('_gemma', {}).get('urgencia', 'unknown') for l in leads if l.get('_gemma_ok')
        )),
        'top_teses': dict(Counter(
            l.get('_gemma', {}).get('tese_juridica_curta', '') for l in leads if l.get('_gemma_ok')
        ).most_common(20))
    }
    return s


async def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--workers', type=int, default=6)
    ap.add_argument('--limit', type=int, default=0, help='0=todos, ou max leads (debug)')
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    log.info("=" * 60)
    log.info("BURNER PREVIDENCIÁRIO MARCO CARPES — INICIANDO")
    log.info("=" * 60)

    # Etapa 1: baixar 12 meses (URLs no JSON ao lado do script)
    script_dir = Path(__file__).resolve().parent
    urls_file = script_dir / 'urls_12meses.json'
    if not urls_file.exists():
        # fallbacks pra outros layouts conhecidos
        for cand in [
            Path('/home/manusalt13/tbr_nero/tools/aurora/marco/urls_12meses.json'),
            Path('/home/manusalt13/tbr_nero/tools/aurora/urls_12meses.json'),
            script_dir.parent / 'urls_12meses.json',
        ]:
            if cand.exists():
                urls_file = cand
                break
    if not urls_file.exists():
        log.error(f"❌ urls_12meses.json não encontrado. Procurei em {script_dir} e fallbacks.")
        sys.exit(1)
    log.info(f"📋 URLs file: {urls_file}")
    files = baixar_meses(urls_file)
    log.info(f"📁 {len(files)} arquivos prontos")

    # Etapa 2: filtrar região Marco
    leads = filtrar_alvo(files)
    if args.limit:
        leads = leads[:args.limit]
        log.info(f"🔬 LIMIT debug: {len(leads):,} leads")

    # Etapa 3: saturar L4
    leads = await saturate_l4(leads, workers=args.workers)

    # Etapa 4: salvar
    out_file = WORK_DIR / f"leads_prev_marco_classificados_{datetime.now():%Y%m%dT%H%M%S}.jsonl"
    to_jsonl(leads, out_file)

    # Etapa 5: stats
    stats = stats_summary(leads)
    stats_file = WORK_DIR / "stats_marco.json"
    with open(stats_file, 'w') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    log.info(f"📊 Stats: {stats_file}")
    log.info(f"  Total: {stats['total']:,}")
    log.info(f"  Score 90+: {stats['score_distribution']['altissimo_90+']:,}")
    log.info(f"  Score 75-89: {stats['score_distribution']['alto_75_89']:,}")
    log.info(f"  Score 50-74: {stats['score_distribution']['medio_50_74']:,}")

    # Etapa 6: upload GCS
    if not args.dry_run:
        upload_gcs(out_file)
        upload_gcs(stats_file)

    log.info("✅ COMPLETO. L4 cumpriu missão Marco Carpes.")


if __name__ == "__main__":
    asyncio.run(main())
